from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field, ConfigDict
from datetime import datetime
import io

from services.auth import verify_token
from services.firestore import get_entries_by_date_range

router = APIRouter(tags=["reports"])


class ReportRequest(BaseModel):
    model_config = ConfigDict(populate_by_name=True)

    from_: str = Field(alias="from")
    to: str


def _parse_range(req: ReportRequest) -> tuple[datetime, datetime]:
    try:
        start = datetime.fromisoformat(req.from_)
        end = datetime.fromisoformat(req.to + "T23:59:59")
        return start, end
    except ValueError:
        raise HTTPException(400, "Invalid date format. Expected YYYY-MM-DD")


@router.post("/excel")
def export_excel(req: ReportRequest, uid: str = Depends(verify_token)):
    import pandas as pd
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    start, end = _parse_range(req)
    entries = get_entries_by_date_range(uid, start, end)

    if not entries:
        raise HTTPException(404, "No entries found for the given date range")

    df = pd.DataFrame(entries)
    df["date_str"] = pd.to_datetime(df["date"]).dt.strftime("%Y-%m-%d")
    df["month"] = pd.to_datetime(df["date"]).dt.to_period("M").astype(str)

    wb = openpyxl.Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")

    def _style_header_row(ws):
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    def _autofit(ws):
        for col in ws.columns:
            width = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(width + 2, 35)

    # Sheet 1: Transactions
    ws1 = wb.active
    ws1.title = "Transactions"
    cols = ["date_str", "type", "amount", "currency", "category", "description"]
    headers = ["Date", "Type", "Amount", "Currency", "Category", "Description"]
    available = [(h, c) for h, c in zip(headers, cols) if c in df.columns]
    ws1.append([h for h, _ in available])
    _style_header_row(ws1)
    for _, row in df[[c for _, c in available]].iterrows():
        ws1.append(list(row))
    _autofit(ws1)

    # Sheet 2: Summary
    ws2 = wb.create_sheet("Summary")
    total_income = float(df[df["type"] == "income"]["amount"].sum())
    total_expenses = float(df[df["type"] == "expense"]["amount"].sum())
    net = total_income - total_expenses
    rate = round(net / total_income * 100, 1) if total_income > 0 else 0.0

    ws2.append(["Metric", "Value"])
    _style_header_row(ws2)
    ws2.append(["Period", f"{req.from_} to {req.to}"])
    ws2.append(["Total Income", round(total_income, 2)])
    ws2.append(["Total Expenses", round(total_expenses, 2)])
    ws2.append(["Net Savings", round(net, 2)])
    ws2.append(["Savings Rate", f"{rate}%"])
    ws2.append(["Total Entries", len(df)])
    _autofit(ws2)

    # Sheet 3: By Category
    ws3 = wb.create_sheet("By Category")
    ws3.append(["Category", "Total", "Count", "Average"])
    _style_header_row(ws3)
    cat_stats = (
        df[df["type"] == "expense"]
        .groupby("category")["amount"]
        .agg(["sum", "count", "mean"])
        .sort_values("sum", ascending=False)
    )
    for cat, row in cat_stats.iterrows():
        ws3.append([cat, round(float(row["sum"]), 2), int(row["count"]), round(float(row["mean"]), 2)])
    _autofit(ws3)

    # Sheet 4: Monthly
    ws4 = wb.create_sheet("Monthly")
    ws4.append(["Month", "Income", "Expenses", "Net Savings"])
    _style_header_row(ws4)
    m_income = df[df["type"] == "income"].groupby("month")["amount"].sum()
    m_expenses = df[df["type"] == "expense"].groupby("month")["amount"].sum()
    for m in sorted(set(m_income.index.tolist()) | set(m_expenses.index.tolist())):
        inc = float(m_income.get(m, 0))
        exp = float(m_expenses.get(m, 0))
        ws4.append([m, round(inc, 2), round(exp, 2), round(inc - exp, 2)])
    _autofit(ws4)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"pocket_report_{req.from_}_to_{req.to}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.post("/tax-summary")
def tax_summary(req: ReportRequest, uid: str = Depends(verify_token)):
    import pandas as pd

    start, end = _parse_range(req)
    entries = get_entries_by_date_range(uid, start, end)

    if not entries:
        return {"period": {"from": req.from_, "to": req.to}, "income": {"total": 0, "byCategory": {}}, "expenses": {"total": 0, "byCategory": {}}, "netIncome": 0}

    df = pd.DataFrame(entries)
    income_by_cat = df[df["type"] == "income"].groupby("category")["amount"].sum()
    expense_by_cat = df[df["type"] == "expense"].groupby("category")["amount"].sum()
    total_income = float(df[df["type"] == "income"]["amount"].sum())
    total_expenses = float(df[df["type"] == "expense"]["amount"].sum())

    return {
        "period": {"from": req.from_, "to": req.to},
        "income": {
            "total": round(total_income, 2),
            "byCategory": {k: round(float(v), 2) for k, v in income_by_cat.items()},
        },
        "expenses": {
            "total": round(total_expenses, 2),
            "byCategory": {k: round(float(v), 2) for k, v in expense_by_cat.items()},
        },
        "netIncome": round(total_income - total_expenses, 2),
    }
